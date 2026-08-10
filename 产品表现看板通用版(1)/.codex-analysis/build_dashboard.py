import json
import math
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据源"
TEMPLATE = Path(__file__).with_name("dashboard_template.html")
OUTPUT_DIR = ROOT / "outputs"
OUTPUT = OUTPUT_DIR / "产品表现离线看板.html"


def find_latest(patterns, required=True):
    files = []
    for pattern in patterns:
        files.extend(ROOT.rglob(pattern))
    files = [p for p in files if p.is_file() and not p.name.startswith("~$")]
    if not files:
        if required:
            raise RuntimeError(f"未找到文件: {patterns}")
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


SOURCE = find_latest(["产品表现ASIN*.xlsx", "*产品表现*ASIN*.xlsx"])
PRODUCTS = find_latest(["售卖产品-数据源*.xlsx", "*售卖产品*数据源*.xlsx"])
BD_FILE = find_latest(["BD活动表*.xlsx", "*BD*活动*.xlsx"], required=False)


def number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text or text in {"-", "--", "N/A", "#N/A"}:
        return 0.0
    if text.endswith("%"):
        text = text[:-1]
        try:
            return float(text) / 100
        except ValueError:
            return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else 0.0


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def clean_date(value):
    text = clean_text(value)
    return text[:10] if text else ""


def norm(value):
    return re.sub(r"[\s\n\r\t（）()_\-—:/：/]+", "", clean_text(value)).casefold()


def read_rows(ws, min_row=1):
    return ws.iter_rows(min_row=min_row, values_only=True)


def first_non_empty_sheet(workbook):
    for ws in workbook.worksheets:
        if ws.max_row > 1 and ws.max_column > 1:
            return ws
    return workbook.worksheets[0]


def find_header_row(ws, required_aliases, scan_rows=10):
    required = [[norm(x) for x in aliases] for aliases in required_aliases]
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=min(scan_rows, ws.max_row), values_only=True), start=1):
        headers = [norm(x) for x in row]
        ok = True
        for aliases in required:
            if not any(alias in headers for alias in aliases):
                ok = False
                break
        if ok:
            return row_num, [clean_text(x) for x in row]
    raise RuntimeError(f"未识别到表头，工作表: {ws.title}")


def col_by_alias(headers, aliases, required=True):
    normalized = [norm(x) for x in headers]
    alias_norms = [norm(x) for x in aliases]
    for alias in alias_norms:
        for i, header in enumerate(normalized):
            if header == alias:
                return i
    for alias in alias_norms:
        for i, header in enumerate(normalized):
            if alias and alias in header:
                return i
    if required:
        raise RuntimeError(f"缺少字段: {aliases}")
    return None


def load_product_mapping():
    wb = openpyxl.load_workbook(PRODUCTS, read_only=True, data_only=True, keep_links=False)
    ws = first_non_empty_sheet(wb)
    header_row, headers = find_header_row(ws, [["ASIN"], ["名称"]])
    cols = {
        "asin": col_by_alias(headers, ["ASIN"]),
        "parent": col_by_alias(headers, ["名称", "父体名称", "父体"]),
        "brand": col_by_alias(headers, ["品牌"], required=False),
        "series": col_by_alias(headers, ["负责人", "系列"], required=False),
        "stage": col_by_alias(headers, ["阶段定位", "阶段"], required=False),
        "ownerStore": col_by_alias(headers, ["店铺", "归属店铺"], required=False),
    }
    mapping = {}
    duplicate_asins = set()
    for row in read_rows(ws, header_row + 1):
        asin = clean_text(row[cols["asin"]] if len(row) > cols["asin"] else "")
        if not asin:
            continue
        if asin in mapping:
            duplicate_asins.add(asin)
        mapping[asin] = {
            "parent": clean_text(row[cols["parent"]] if len(row) > cols["parent"] else ""),
            "brand": clean_text(row[cols["brand"]] if cols["brand"] is not None and len(row) > cols["brand"] else ""),
            "series": clean_text(row[cols["series"]] if cols["series"] is not None and len(row) > cols["series"] else ""),
            "stage": clean_text(row[cols["stage"]] if cols["stage"] is not None and len(row) > cols["stage"] else ""),
            "ownerStore": clean_text(row[cols["ownerStore"]] if cols["ownerStore"] is not None and len(row) > cols["ownerStore"] else ""),
        }
    wb.close()
    return mapping, sorted(duplicate_asins)


def load_bd_intervals_by_parent():
    if not BD_FILE:
        return [], {}

    wb = openpyxl.load_workbook(BD_FILE, read_only=True, data_only=True, keep_links=False)
    ws = first_non_empty_sheet(wb)
    top_rows = list(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True))
    intervals = []

    # Current BD source is grouped horizontally:
    # row 1 = parent name, row 2 = start/end labels, rows below = date pairs.
    for header_row_index in range(min(5, len(top_rows))):
        row = top_rows[header_row_index]
        for col, value in enumerate(row):
            if "开始" not in clean_text(value):
                continue
            end_col = None
            for candidate in range(col + 1, min(col + 4, ws.max_column)):
                if "结束" in clean_text(ws.cell(header_row_index + 1, candidate + 1).value):
                    end_col = candidate
                    break
            if end_col is None:
                continue

            parent = ""
            for parent_row_index in range(header_row_index - 1, -1, -1):
                for parent_col in range(col, -1, -1):
                    parent = clean_text(top_rows[parent_row_index][parent_col] if parent_col < len(top_rows[parent_row_index]) else "")
                    if parent:
                        break
                if parent:
                    break
            if not parent:
                continue

            for row_num in range(header_row_index + 2, ws.max_row + 1):
                start = clean_date(ws.cell(row_num, col + 1).value)
                end = clean_date(ws.cell(row_num, end_col + 1).value)
                if not start or not end:
                    continue
                if end < start:
                    start, end = end, start
                intervals.append({"parent": parent, "start": start, "end": end})

    wb.close()
    by_parent = {}
    unique = []
    seen = set()
    for interval in intervals:
        key = (interval["parent"], interval["start"], interval["end"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(interval)
        by_parent.setdefault(interval["parent"], []).append(interval)
    return unique, by_parent


def is_bd_day(parent, day, bd_by_parent):
    return any(item["start"] <= day <= item["end"] for item in bd_by_parent.get(parent, []))


def load_source_rows(mapping, bd_by_parent):
    source_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True, keep_links=False)
    ws = first_non_empty_sheet(source_wb)
    header_row, headers = find_header_row(ws, [["日期"], ["ASIN"], ["店铺"], ["国家"]])
    cols = {
        "date": col_by_alias(headers, ["日期"]),
        "asin": col_by_alias(headers, ["ASIN"]),
        "store": col_by_alias(headers, ["店铺"]),
        "country": col_by_alias(headers, ["国家"]),
        "units": col_by_alias(headers, ["销量"]),
        "sales": col_by_alias(headers, ["销售额"]),
        "orders": col_by_alias(headers, ["订单量"]),
        "netSales": col_by_alias(headers, ["净销售额"]),
        "b2bUnits": col_by_alias(headers, ["B2B 销量", "B2B销量"]),
        "b2bOrders": col_by_alias(headers, ["B2B 订单量", "B2B订单量"]),
        "profit": col_by_alias(headers, ["订单毛利润"]),
        "returns": col_by_alias(headers, ["退货量"]),
        "sessionsBrowser": col_by_alias(headers, ["Sessions-Browser"]),
        "sessionsMobile": col_by_alias(headers, ["Sessions-Mobile"]),
        "sessions": col_by_alias(headers, ["Sessions-Total"]),
        "impressions": col_by_alias(headers, ["展示"]),
        "clicks": col_by_alias(headers, ["点击"]),
        "spend": col_by_alias(headers, ["广告花费"]),
        "adSales": col_by_alias(headers, ["广告销售额"]),
        "adOrders": col_by_alias(headers, ["广告订单量"]),
        "naturalClicks": col_by_alias(headers, ["自然点击量"]),
        "naturalOrders": col_by_alias(headers, ["自然订单量"]),
    }

    compact_rows = []
    source_count = 0
    source_asins = set()
    matched_bd_parents = set()
    matched_bd_intervals = set()

    for row in read_rows(ws, header_row + 1):
        source_count += 1
        asin = clean_text(row[cols["asin"]] if len(row) > cols["asin"] else "")
        day = clean_date(row[cols["date"]] if len(row) > cols["date"] else "")
        if asin:
            source_asins.add(asin)
        m = mapping.get(asin)
        if not m:
            continue

        parent = m["parent"] or "未分类"
        bd_flag = 1 if day and is_bd_day(parent, day, bd_by_parent) else 0
        if bd_flag:
            matched_bd_parents.add(parent)
            for interval in bd_by_parent.get(parent, []):
                if interval["start"] <= day <= interval["end"]:
                    matched_bd_intervals.add((parent, interval["start"], interval["end"]))

        def v(key):
            col = cols[key]
            return row[col] if len(row) > col else None

        compact_rows.append({
            "date": day,
            "asin": asin,
            "store": clean_text(v("store")),
            "country": clean_text(v("country")),
            "stage": m["stage"] or "未分类",
            "series": m["series"] or "未分类",
            "parent": parent,
            "ownerStore": m["ownerStore"] or "未分类",
            "bd": bd_flag,
            "units": round(number(v("units")), 4),
            "sales": round(number(v("sales")), 4),
            "orders": round(number(v("orders")), 4),
            "netSales": round(number(v("netSales")), 4),
            "b2bUnits": round(number(v("b2bUnits")), 4),
            "b2bOrders": round(number(v("b2bOrders")), 4),
            "profit": round(number(v("profit")), 4),
            "returns": round(number(v("returns")), 4),
            "sessionsBrowser": round(number(v("sessionsBrowser")), 4),
            "sessionsMobile": round(number(v("sessionsMobile")), 4),
            "sessions": round(number(v("sessions")), 4),
            "impressions": round(number(v("impressions")), 4),
            "clicks": round(number(v("clicks")), 4),
            "spend": round(number(v("spend")), 4),
            "adSales": round(number(v("adSales")), 4),
            "adOrders": round(number(v("adOrders")), 4),
            "naturalClicks": round(number(v("naturalClicks")), 4),
            "naturalOrders": round(number(v("naturalOrders")), 4),
        })
    source_wb.close()
    return compact_rows, {
        "sourceRows": source_count,
        "sourceAsins": source_asins,
        "matchedBdParents": matched_bd_parents,
        "matchedBdIntervals": matched_bd_intervals,
    }


def main():
    mapping, duplicate_mapping_asins = load_product_mapping()
    bd_intervals, bd_by_parent = load_bd_intervals_by_parent()
    compact_rows, source_meta = load_source_rows(mapping, bd_by_parent)
    if not compact_rows:
        raise RuntimeError("没有映射到任何售卖产品 ASIN，请检查售卖产品-数据源。")

    source_asins = source_meta["sourceAsins"]
    mapping_asins = set(mapping)
    source_parents = {mapping[asin]["parent"] for asin in source_asins & mapping_asins if mapping[asin]["parent"]}
    bd_parents = set(bd_by_parent)
    dates = sorted({r["date"] for r in compact_rows if r["date"]})
    overlapped_intervals = [
        interval for interval in bd_intervals
        if interval["parent"] in source_parents and dates and interval["end"] >= dates[0] and interval["start"] <= dates[-1]
    ]
    options = {
        key: sorted({r[key] for r in compact_rows}, key=lambda x: x.casefold())
        for key in ["store", "country", "stage", "series", "parent"]
    }
    bd_schedule_by_parent = {
        parent: [
            f"{item['start']} 至 {item['end']}"
            for item in sorted(items, key=lambda x: (x["start"], x["end"]))
        ]
        for parent, items in bd_by_parent.items()
    }

    audit = {
        "sourceOnlyAsins": sorted(source_asins - mapping_asins),
        "mappingOnlyAsins": sorted(mapping_asins - source_asins),
        "duplicateMappingAsins": duplicate_mapping_asins,
        "bdParentsWithoutProduct": sorted(bd_parents - {m["parent"] for m in mapping.values() if m["parent"]}),
        "productParentsWithoutBd": sorted({m["parent"] for m in mapping.values() if m["parent"]} - bd_parents),
    }
    audit_counts = {key: len(value) for key, value in audit.items()}

    data = {
        "meta": {
            "sourceFile": SOURCE.name,
            "productFile": PRODUCTS.name,
            "bdFile": BD_FILE.name if BD_FILE else "",
            "minDate": dates[0],
            "maxDate": dates[-1],
            "sourceRows": source_meta["sourceRows"],
            "sourceAsins": len(source_asins),
            "mappedRows": len(compact_rows),
            "mappedAsins": len({r["asin"] for r in compact_rows}),
            "mappingAsins": len(mapping),
            "bdIntervals": len(bd_intervals),
            "bdOverlapIntervals": len(overlapped_intervals),
            "bdMatchedIntervals": len(source_meta["matchedBdIntervals"]),
            "bdMatchedParents": len(source_meta["matchedBdParents"]),
            "auditCounts": audit_counts,
            "auditSamples": {key: value[:20] for key, value in audit.items()},
            "options": options,
            "bdScheduleByParent": bd_schedule_by_parent,
        },
        "rows": compact_rows,
    }
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    html = TEMPLATE.read_text(encoding="utf-8").replace("__DATA_JSON__", payload)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(html, encoding="utf-8")

    totals = {
        key: round(sum(r[key] for r in compact_rows), 4)
        for key in [
            "units", "orders", "netSales", "profit", "spend", "adSales",
            "sessions", "adOrders", "naturalOrders", "returns",
        ]
    }
    totals["tacos"] = round(totals["spend"] / totals["netSales"], 8) if totals["netSales"] else None
    totals["margin"] = round(totals["profit"] / totals["netSales"], 8) if totals["netSales"] else None
    totals["orderCvr"] = round(totals["orders"] / totals["sessions"], 8) if totals["sessions"] else None
    stats = {
        "output": str(OUTPUT),
        "bytes": OUTPUT.stat().st_size,
        "sourceRows": source_meta["sourceRows"],
        "mappedRows": len(compact_rows),
        "mappedAsins": len({r["asin"] for r in compact_rows}),
        "dateRange": [dates[0], dates[-1]],
        "options": {k: len(v) for k, v in options.items()},
        "bd": {
            "bdFile": BD_FILE.name if BD_FILE else "",
            "intervals": len(bd_intervals),
            "overlapIntervals": len(overlapped_intervals),
            "matchedIntervals": len(source_meta["matchedBdIntervals"]),
            "matchedParents": len(source_meta["matchedBdParents"]),
            "bdRows": sum(r["bd"] for r in compact_rows),
        },
        "auditCounts": audit_counts,
        "totals": totals,
    }
    Path(__file__).with_name("dashboard_stats.json").write_text(
        json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
